#!/usr/bin/env python3
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors
import subprocess

def write_excel(filename, url, asking, units, ave_rent, sqr_ft, vac_rate, other_in, \
ex_1, ex_2, ex_3, ex_4, ex_5, ex_6, ex_7, ex_8, ex_9, ex_10, ex_11):
    workbook = load_workbook(filename='Template2_0.xlsx')
    sheet = workbook.active
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    #url
    sheet['A2'] = url
    #asking price
    sheet['B4'] = asking
    #number of units
    sheet['B12'] = units
    #ave monthly rent
    sheet['B13'] = ave_rent
    #total sqr-ft
    sheet['N4'] = sqr_ft
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    #vacancy rate
    sheet['B22'] = vac_rate
    #other income
    sheet['F24'] = other_in
    #expenses
    sheet['F28'] = ex_1
    sheet['F29'] = ex_2
    sheet['F30'] = ex_3
    sheet['F31'] = ex_4
    sheet['F32'] = ex_5
    sheet['F33'] = ex_6
    sheet['F34'] = ex_7
    sheet['F35'] = ex_8
    sheet['F36'] = ex_9
    sheet['F37'] = ex_10
    sheet['F38'] = ex_11
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # Save the spreadsheet
    workbook.save(filename='../deals/'+filename)

#~~~~~~~~~~~~~~~~~current~~affairs~~~~~~~~~~~~~~~~~~~~~
url = 'https://www.crexi.com/properties/317972/minnesota-241-w-lake-ave'
write_excel('current_test.xlsx', url, 240000, 12,500,10000,.1,0,60,0,0,0,0,0,0,0,0,1850,30000)
subprocess.run(['open', '../deals/current_test.xlsx'])
print('finished.')