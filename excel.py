#!/usr/bin/env python3
#
# import sys
# import os
# import csv
# from PyQt5 import uic
# from PyQt5.QtWidgets import QTableWidget, QApplication, QMainWindow, QTableWidgetItem, QFileDialog
#
# analysis_view, QtBaseClass = uic.loadUiType('guis/analysis.ui')
# class MyTable(QTableWidget):
#     def __init__(self, r, c):
#         super().__init__(r, c)
#         self.check_change = True
#         self.init_ui()
#
#     def init_ui(self):
#         self.cellChanged.connect(self.c_current)
#         self.show()
#
#     def c_current(self):
#         if self.check_change:
#             row = self.currentRow()
#             col = self.currentColumn()
#             value = self.item(row, col)
#             value = value.text()
#             print("The current cell is ", row, ", ", col)
#             print("In this cell we have: ", value)
#
#     def open_sheet(self):
#         self.check_change = False
#         path = QFileDialog.getOpenFileName(self, 'Open CSV', os.getenv('/home/tetrapro/projects/python/Rightside_Two.0'), 'CSV(*.csv)')
#         if path[0] != '':
#             with open(path[0], newline='') as csv_file:
#                 self.setRowCount(0)
#                 self.setColumnCount(10)
#                 my_file = csv.reader(csv_file, delimiter=',', quotechar='|')
#                 for row_data in my_file:
#                     row = self.rowCount()
#                     self.insertRow(row)
#                     if len(row_data) > 10:
#                         self.setColumnCount(len(row_data))
#                     for column, stuff in enumerate(row_data):
#                         item = QTableWidgetItem(stuff)
#                         self.setItem(row, column, item)
#         self.check_change = True
#     def save_sheet(self):
#         path = QFileDialog.getOpenFileName(self, 'Save CSV', os.getenv('/home/tetrapro/projects/python/Rightside_Two.0'), 'CSV(*.csv)')
#         if path[0] != '':
#             with open(path[0], 'w') as csv_file:
#                 writer = csv.writer(csv_file, dialect='excel')
#                 for row in range(self.rowCount()):
#                     row_data = []
#                     for column in range(self.columnCount()):
#                         item = self.item(row, column)
#                         if item is not None:
#                             row_data.append(item.next())
#                         else:
#                             row_data.append('')
#                     writer.writerow(row_data)
# class Sheet(QMainWindow):
#     def __init__(self):
#         super().__init__()
#
#         self.form_widget = MyTable(10, 10)
#         self.setCentralWidget(self.form_widget)
#         col_headers = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
#         self.form_widget.setHorizontalHeaderLabels(col_headers)
#
#         self.form_widget.open_sheet()
#
#         self.show()
#
# app = QApplication(sys.argv)
# sheet = Sheet()
# sys.exit(app.exec_())

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors
import pandas as pd

def write_excel(filename, url, asking, units, ave_rent, sqr_ft, vac_rate, other_in, \
ex_1, ex_2, ex_3, ex_4, ex_5, ex_6, ex_7, ex_8, ex_9, ex_10, ex_11):
    workbook = load_workbook(filename='Template2_0.xlsx')
    sheet = workbook.active
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    #url
    sheet['A2'] = url
    #asking price
    sheet['B4'] = asking
    #number of units
    sheet['B12'] = units
    #ave monthly rent
    sheet['B13'] = ave_rent
    #total sqr-ft
    sheet['N4'] = sqr_ft
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    #vacancy rate
    sheet['B22'] = vac_rate
    #other income
    sheet['F24'] = other_in
    #expenses
    sheet['F28'] = ex_1
    sheet['F29'] = ex_2
    sheet['F30'] = ex_3
    sheet['F31'] = ex_4
    sheet['F32'] = ex_5
    sheet['F33'] = ex_6
    sheet['F34'] = ex_7
    sheet['F35'] = ex_8
    sheet['F36'] = ex_9
    sheet['F37'] = ex_10
    sheet['F38'] = ex_11
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # Save the spreadsheet
    workbook.save(filename='deals/'+filename)

#~~~~~~~~~~~~~~~~~current~~affairs~~~~~~~~~~~~~~~~~~~~~
url = 'https://www.crexi.com/properties/317972/minnesota-241-w-lake-ave'
write_excel('current.xlsx', url, 240000, 12,500,10000,.1,0,60,0,0,0,0,0,0,0,0,1850,30000)
read_file = pd.read_excel(r'deals/current.xlsx')
read_file.to_csv(r'deals/current.csv')
#~~~~~~~~~~~~~~~~Rightside_Two~~~~~~~~~~~~~~~~~~~~~~~~
url = 'https://www.crexi.com/properties/317972/minnesota-241-w-lake-ave'
write_excel('rightside.xlsx', url, 235000, 12,525,10000,.1,1500,60,0,0,0,0,0,0,0,0,1850,30000)
read_file = pd.read_excel(r'deals/rightside.xlsx')
read_file.to_csv(r'deals/rightside.csv')
